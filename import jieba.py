import jieba.analyse
import openpyxl
import re
import os

# -------------------------- 配置项（请根据实际情况修改） --------------------------
STOPWORDS_PATH = "D:\\PYECourse\\stopwords.txt"  # 停用词表路径
DEFAULT_FILE_PATH = "D:\\PYECourse\\paper.txt"    # 默认文献文本路径
NAME_PINYIN = "zhangsan"                          # 姓名拼音（用于Excel文件名）
# --------------------------------------------------------------------------------

def main():   
    """程序主入口，串联整个科研小助手执行流程"""
    # 1. 获取文件路径（支持用户输入或使用默认路径）
    print("===== 科研小助手 =====")
    file_path = input(f"请输入文献文件路径（默认：{DEFAULT_FILE_PATH}）：").strip()
    if not file_path:
        file_path = DEFAULT_FILE_PATH
    
    # 2. 读取并预处理笔记
    try:
        clean_content = read_and_preprocess_notes(file_path)
        if not clean_content:
            print("错误：预处理后无有效文本内容")
            return
    except Exception as e:
        print(f"读取/预处理文件失败：{e}")
        return
    
    # 3. 提取关键词（TF-IDF算法，前10个）
    try:
        keywords = get_keywords(clean_content)
        print(f"成功提取核心关键词：{keywords}")
    except Exception as e:
        print(f"提取关键词失败：{e}")
        return
    
    # 4. 保存结果到Excel
    try:
        excel_path = save_result(clean_content, keywords)
        print(f"结果已保存至Excel：{excel_path}")
    except Exception as e:
        print(f"保存Excel失败：{e}")
        return

def load_stopwords():
    """加载停用词表，返回停用词集合（辅助函数）"""
    stopwords = set()
    # 1. 加载外部停用词表
    if os.path.exists(STOPWORDS_PATH):
        try:
            with open(STOPWORDS_PATH, "r", encoding="utf-8") as f:
                for line in f:
                    word = line.strip()
                    if word:
                        stopwords.add(word)
            print(f"成功加载停用词表，共{len(stopwords)}个停用词")
        except Exception as e:
            print(f"加载停用词表失败，使用默认停用词：{e}")
            stopwords = {"是", "的", "等", "与", "也", "这", "在", "和"}
    else:
        print("未找到停用词表，使用默认停用词")
        stopwords = {"是", "的", "等", "与", "也", "这", "在", "和"}
    return stopwords

def read_and_preprocess_notes(file_path):
    """读取笔记内容并进行预处理（核心封装函数）"""
    # 第一步：识别文件类型（后缀+文件头双重验证）
    file_type = identify_file_type(file_path)
    if not file_type:
        raise ValueError("文件类型不匹配，请检查文件格式")
    
    # 第二步：根据文件类型读取原始内容
    raw_content = read_file_by_type(file_path, file_type)
    if not raw_content:
        raise ValueError("读取到的文件内容为空")
    
    # 第三步：清理乱码和无意义字符
    clean_content = clean_invalid_chars(raw_content, file_type)
    return clean_content

def identify_file_type(file_path):
    """识别文件类型：结合后缀名和文件头特征"""
    # 校验文件是否存在
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在：{file_path}")
    
    # 提取后缀
    suffix = os.path.splitext(file_path)[-1].lower().lstrip(".")
    if suffix not in ["txt", "md", "docx", "pdf"]:
        print(f"警告：不推荐的文件类型{suffix}，将尝试按txt读取")
        suffix = "txt"
    
    # 读取文件头验证（简化版，适配核心场景）
    with open(file_path, "rb") as f:
        file_header = f.read(16)
    
    # 验证规则
    if suffix == "txt" or suffix == "md":
        return suffix
    elif suffix == "docx" and file_header.startswith(b"\x50\x4b\x03\x04"):
        return "docx"
    elif suffix == "pdf" and file_header.startswith(b"%PDF-"):
        return "pdf"
    else:
        print("文件后缀与实际类型不匹配，返回None")
        return None

def read_file_by_type(file_path, file_type):
    """按文件类型分支读取内容"""
    if file_type == "txt":
        return read_txt_file(file_path)
    elif file_type == "md":
        return read_md_file(file_path)
    elif file_type == "docx":
        return read_docx_file(file_path)
    elif file_type == "pdf":
        return read_pdf_file(file_path)
    else:
        raise ValueError(f"不支持的文件类型：{file_type}")

def read_txt_file(file_path):
    """读取txt文件，处理多编码兼容"""
    encodings = ["utf-8", "gbk", "utf-8-sig", "gb2312"]
    for encoding in encodings:
        try:
            with open(file_path, "r", encoding=encoding) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
    raise ValueError(f"txt文件{file_path}编码无法识别")

def read_md_file(file_path):
    """读取md文件，复用txt逻辑"""
    return read_txt_file(file_path)

def read_docx_file(file_path):
    """读取docx文件（需安装python-docx：pip install python-docx）"""
    try:
        from docx import Document
        doc = Document(file_path)
        content = []
        for para in doc.paragraphs:
            content.append(para.text)
        return "\n".join(content)
    except ImportError:
        raise ImportError("请安装python-docx库：pip install python-docx")
    except Exception as e:
        raise ValueError(f"读取docx失败：{e}")

def read_pdf_file(file_path):
    """读取pdf文件（需安装PyPDF2：pip install PyPDF2）"""
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(file_path)
        content = []
        for page in reader.pages:
            text = page.extract_text() or ""
            content.append(text)
        return "\n".join(content)
    except ImportError:
        raise ImportError("请安装PyPDF2库：pip install PyPDF2")
    except Exception as e:
        raise ValueError(f"读取pdf失败：{e}")

def clean_invalid_chars(content, file_type):
    """清理乱码、无意义字符"""
    # 通用清理：移除控制字符、乱码、多余空格
    clean_content = re.sub(r"[\x00-\x1f\x7f-\xff]", "", content)  # 移除控制字符/乱码
    clean_content = re.sub(r"\s+", " ", clean_content).strip()   # 合并多余空格
    
    # 专属清理：md文件移除标签，pdf/docx移除特殊符号
    if file_type == "md":
        clean_content = re.sub(r"[#*`\[\]\(\)]|!\[.*?\]\(.*?\)", "", clean_content)
    elif file_type in ["docx", "pdf"]:
        clean_content = re.sub(r"[^\u4e00-\u9fa5a-zA-Z0-9，。！？；：,.!?;:\s]", "", clean_content)
    
    return clean_content

def get_keywords(content):
    """从清理后的内容提取前10个核心关键词（TF-IDF算法+停用词过滤）"""
    # 加载停用词
    stopwords = load_stopwords()
    
    # 配置jieba TF-IDF参数：停用词+返回前10个关键词
    jieba.analyse.set_stop_words(STOPWORDS_PATH)  # 绑定停用词表
    # extract_tags参数：文本、返回数量、是否带权重、停用词集合
    keywords = jieba.analyse.extract_tags(
        content,
        topK=10,
        withWeight=False,
        allowPOS=()  # 不限制词性，如需精准可设为["n","v"]（名词/动词）
    )
    
    # 二次过滤：确保无单个字、无空字符串（兜底）
    filtered_keywords = [word for word in keywords if len(word.strip())>1 and word not in stopwords]
    # 若过滤后不足10个，补充至10个（避免空值）
    if len(filtered_keywords) < 10:
        filtered_keywords += [""] * (10 - len(filtered_keywords))
    
    return filtered_keywords[:10]  # 确保只返回10个

def save_result(content, keywords):
    """保存文献段落+关键词到Excel（openpyxl）"""
    # 1. 新建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "科研关键词提取结果"
    
    # 2. 写入表头
    ws["A1"] = "文献段落"
    ws["B1"] = "核心关键词（前10）"
    
    # 3. 写入内容：A2=文献段落，B2=关键词（逗号分隔）
    ws["A2"] = content[:50000]  # Excel单元格最大字符数限制（避免溢出）
    ws["B2"] = "，".join(keywords)
    
    # 4. 调整列宽（优化显示）
    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 50
    
    # 5. 保存文件
    excel_filename = f"{NAME_PINYIN}_research.xlsx"
    excel_path = os.path.join(os.path.dirname(DEFAULT_FILE_PATH), excel_filename)
    wb.save(excel_path)
    
    return excel_path

# 程序执行入口
if __name__ == "__main__":
    main()